from __future__ import annotations

import csv
import math
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from pathlib import Path
from typing import Dict, List, Optional


@dataclass
class LabelStats:
    class_id: int
    instance_count: int = 0
    image_count: int = 0


@dataclass
class FolderStats:
    folder: str
    total_images: int
    labeled_images: int
    empty_images: int
    total_boxes: int
    avg_boxes_per_image: float
    small_count: int
    medium_count: int
    large_count: int
    area_bins: Dict[str, int]
    aspect_ratio_bins: Dict[str, int]
    class_stats: List[LabelStats]
    image_box_counts: List[int]
    box_areas: List[float]
    box_aspect_ratios: List[float]


def _parse_label_file(path: Path):
    boxes = []
    with path.open("r", encoding="utf-8") as f:
        for line in f:
            parts = line.strip().split()
            if len(parts) < 5:
                continue
            try:
                class_id = int(parts[0])
                x = float(parts[1])
                y = float(parts[2])
                w = float(parts[3])
                h = float(parts[4])
            except ValueError:
                continue
            boxes.append((class_id, x, y, w, h))
    return boxes


def _bin_label(value: float, edges: List[float], labels: List[str]) -> str:
    for idx, edge in enumerate(edges):
        if value < edge:
            return labels[idx]
    return labels[-1]


def analyze_folder(folder_path: str) -> FolderStats:
    folder = Path(folder_path)
    if not folder.is_dir():
        raise ValueError(f"{folder_path!r} is not a valid folder")

    txt_files = sorted(folder.glob("*.txt"))
    total_images = len(txt_files)
    empty_images = 0
    labeled_images = 0

    class_instances = Counter()
    class_images = Counter()
    image_box_counts: List[int] = []
    box_areas: List[float] = []
    box_aspect_ratios: List[float] = []

    small_count = medium_count = large_count = 0
    area_bins = Counter()
    aspect_bins = Counter()
    aspect_edges = [0.5, 1.0, 2.0, 5.0]
    aspect_labels = ["<0.5", "0.5-1", "1-2", "2-5", ">=5"]
    area_edges = [0.01, 0.05, 0.15, 0.30]
    area_labels = ["<1%", "1-5%", "5-15%", "15-30%", ">=30%"]

    for txt in txt_files:
        boxes = _parse_label_file(txt)
        if not boxes:
            empty_images += 1
            image_box_counts.append(0)
            continue

        labeled_images += 1
        image_box_counts.append(len(boxes))

        seen_classes = set()
        for class_id, _, _, w, h in boxes:
            class_instances[class_id] += 1
            seen_classes.add(class_id)

            area = max(0.0, w * h)
            aspect = (w / h) if h > 0 else 0.0
            box_areas.append(area)
            if aspect > 0:
                box_aspect_ratios.append(aspect)

            if area < 0.02:
                small_count += 1
            elif area < 0.12:
                medium_count += 1
            else:
                large_count += 1

            area_bins[_bin_label(area, area_edges, area_labels)] += 1
            if aspect > 0:
                aspect_bins[_bin_label(aspect, aspect_edges, aspect_labels)] += 1

        for cid in seen_classes:
            class_images[cid] += 1

    total_boxes = sum(class_instances.values())
    avg_boxes_per_image = (total_boxes / total_images) if total_images else 0.0

    class_ids = sorted(set(class_instances.keys()) | set(class_images.keys()))
    class_stats = [
        LabelStats(
            class_id=cid,
            instance_count=class_instances.get(cid, 0),
            image_count=class_images.get(cid, 0),
        )
        for cid in class_ids
    ]

    return FolderStats(
        folder=str(folder),
        total_images=total_images,
        labeled_images=labeled_images,
        empty_images=empty_images,
        total_boxes=total_boxes,
        avg_boxes_per_image=avg_boxes_per_image,
        small_count=small_count,
        medium_count=medium_count,
        large_count=large_count,
        area_bins=dict(area_bins),
        aspect_ratio_bins=dict(aspect_bins),
        class_stats=class_stats,
        image_box_counts=image_box_counts,
        box_areas=box_areas,
        box_aspect_ratios=box_aspect_ratios,
    )


def export_csv(stats: FolderStats, path: str) -> None:
    with open(path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(["Summary"])
        writer.writerow(["folder", stats.folder])
        writer.writerow(["total_images", stats.total_images])
        writer.writerow(["labeled_images", stats.labeled_images])
        writer.writerow(["empty_images", stats.empty_images])
        writer.writerow(["total_boxes", stats.total_boxes])
        writer.writerow(["avg_boxes_per_image", f"{stats.avg_boxes_per_image:.4f}"])
        writer.writerow(["small_count", stats.small_count])
        writer.writerow(["medium_count", stats.medium_count])
        writer.writerow(["large_count", stats.large_count])
        writer.writerow([])
        writer.writerow(["class_id", "instance_count", "image_count"])
        for item in stats.class_stats:
            writer.writerow([item.class_id, item.instance_count, item.image_count])
        writer.writerow([])
        writer.writerow(["area_bins"])
        for k, v in stats.area_bins.items():
            writer.writerow([k, v])
        writer.writerow([])
        writer.writerow(["aspect_ratio_bins"])
        for k, v in stats.aspect_ratio_bins.items():
            writer.writerow([k, v])


def export_xlsx(stats: FolderStats, path: str) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("Excel export requires openpyxl") from e

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    ws.append(["folder", stats.folder])
    ws.append(["total_images", stats.total_images])
    ws.append(["labeled_images", stats.labeled_images])
    ws.append(["empty_images", stats.empty_images])
    ws.append(["total_boxes", stats.total_boxes])
    ws.append(["avg_boxes_per_image", stats.avg_boxes_per_image])
    ws.append(["small_count", stats.small_count])
    ws.append(["medium_count", stats.medium_count])
    ws.append(["large_count", stats.large_count])

    ws2 = wb.create_sheet("Classes")
    ws2.append(["class_id", "instance_count", "image_count"])
    for item in stats.class_stats:
        ws2.append([item.class_id, item.instance_count, item.image_count])

    ws3 = wb.create_sheet("Distributions")
    ws3.append(["area_bin", "count"])
    for k, v in stats.area_bins.items():
        ws3.append([k, v])
    ws3.append([])
    ws3.append(["aspect_ratio_bin", "count"])
    for k, v in stats.aspect_ratio_bins.items():
        ws3.append([k, v])

    wb.save(path)
